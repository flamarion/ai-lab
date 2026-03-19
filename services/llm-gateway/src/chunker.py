"""Document loading and chunking for RAG ingestion."""

import logging
import re
from pathlib import Path

logger = logging.getLogger(__name__)

# File extensions we can load
TEXT_EXTENSIONS = {".txt", ".md", ".rst", ".csv", ".log", ".tsv"}
CODE_EXTENSIONS = {
    ".py", ".js", ".ts", ".go", ".rs", ".rb", ".java", ".kt",
    ".c", ".cpp", ".h", ".hpp", ".cs", ".sh", ".bash", ".zsh",
    ".yaml", ".yml", ".toml", ".json", ".xml", ".html", ".css",
    ".sql", ".tf", ".hcl", ".Dockerfile", ".env", ".ini", ".cfg",
    ".r", ".scala", ".lua", ".pl", ".php", ".swift",
}
PDF_EXTENSIONS = {".pdf"}
OFFICE_EXTENSIONS = {".docx", ".xlsx"}

SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | CODE_EXTENSIONS | PDF_EXTENSIONS | OFFICE_EXTENSIONS


def load_file(path: str) -> str:
    """Read a file and return its text content."""
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix in PDF_EXTENSIONS:
        return _load_pdf(path)
    if suffix == ".docx":
        return _load_docx(p.read_bytes())
    if suffix == ".xlsx":
        return _load_xlsx(p.read_bytes())
    if suffix in TEXT_EXTENSIONS | CODE_EXTENSIONS or p.name == "Dockerfile":
        return p.read_text(encoding="utf-8", errors="replace")

    raise ValueError(f"Unsupported file type: {suffix}")


def load_bytes(content: bytes, filename: str) -> str:
    """Load content from bytes (for file uploads)."""
    name = Path(filename).name
    suffix = Path(filename).suffix.lower()

    if suffix in PDF_EXTENSIONS:
        return _load_pdf_bytes(content)
    if suffix == ".docx":
        return _load_docx(content)
    if suffix == ".xlsx":
        return _load_xlsx(content)
    if suffix in TEXT_EXTENSIONS | CODE_EXTENSIONS:
        return content.decode("utf-8", errors="replace")

    # Handle dotfiles (.env) and extensionless files (Dockerfile, Makefile)
    if name in ("Dockerfile", "Makefile", ".env", ".gitignore", ".dockerignore"):
        return content.decode("utf-8", errors="replace")

    # Last resort: try UTF-8 decode for unknown text files
    try:
        return content.decode("utf-8")
    except UnicodeDecodeError:
        raise ValueError(f"Unsupported file type: {suffix or name}")


def _load_pdf(path: str) -> str:
    """Extract text from a PDF file."""
    from pypdf import PdfReader

    reader = PdfReader(path)
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def _load_pdf_bytes(content: bytes) -> str:
    """Extract text from PDF bytes."""
    import io
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n\n".join(pages)


def _load_docx(content: bytes) -> str:
    """Extract text from a .docx file."""
    import io
    from docx import Document

    doc = Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)


def _load_xlsx(content: bytes) -> str:
    """Extract text from an .xlsx spreadsheet (all sheets)."""
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                rows.append("\t".join(cells))
        if rows:
            parts.append(f"[Sheet: {sheet}]\n" + "\n".join(rows))
    wb.close()
    return "\n\n".join(parts)


def chunk_text(
    text: str,
    chunk_size: int = 512,
    overlap: int = 50,
) -> list[dict]:
    """Split text into overlapping chunks.

    Strategy: split on paragraphs first, then sentences, then by character
    count. Each chunk records its position in the original text.

    Args:
        text: The full document text.
        chunk_size: Target chunk size in characters (~tokens * 4).
        overlap: Number of characters to overlap between chunks.

    Returns:
        List of dicts with keys: text, chunk_index, start_offset, end_offset.
    """
    if not text.strip():
        return []

    # Normalize whitespace but preserve paragraph breaks
    text = re.sub(r"\r\n", "\n", text)

    # Split into paragraphs (double newline)
    paragraphs = re.split(r"\n\s*\n", text)
    paragraphs = [p.strip() for p in paragraphs if p.strip()]

    # Build chunks by accumulating paragraphs up to chunk_size
    chunks = []
    current = ""
    current_start = 0
    offset = 0

    for para in paragraphs:
        # Find the actual offset of this paragraph in the original text
        para_start = text.find(para, offset)
        if para_start == -1:
            para_start = offset

        if not current:
            current = para
            current_start = para_start
        elif len(current) + len(para) + 2 <= chunk_size:
            current = current + "\n\n" + para
        else:
            # Flush current chunk
            chunks.append({
                "text": current,
                "chunk_index": len(chunks),
                "start_offset": current_start,
                "end_offset": current_start + len(current),
            })
            # Start new chunk with overlap from the end of the previous
            if overlap > 0 and len(current) > overlap:
                prev_len = len(current)
                overlap_text = current[-overlap:]
                current = overlap_text + "\n\n" + para
                current_start = current_start + prev_len - len(overlap_text)
            else:
                current = para
                current_start = para_start

        offset = para_start + len(para)

    # Don't forget the last chunk
    if current.strip():
        chunks.append({
            "text": current,
            "chunk_index": len(chunks),
            "start_offset": current_start,
            "end_offset": current_start + len(current),
        })

    # If any chunk is still too large, split by sentences
    final_chunks = []
    for chunk in chunks:
        if len(chunk["text"]) <= chunk_size * 1.5:
            chunk["chunk_index"] = len(final_chunks)
            final_chunks.append(chunk)
        else:
            sub_chunks = _split_large_chunk(chunk["text"], chunk_size, chunk["start_offset"])
            for sc in sub_chunks:
                sc["chunk_index"] = len(final_chunks)
                final_chunks.append(sc)

    return final_chunks


def _split_large_chunk(text: str, chunk_size: int, base_offset: int) -> list[dict]:
    """Split an oversized chunk by sentences."""
    sentences = re.split(r"(?<=[.!?])\s+", text)
    chunks = []
    current = ""
    current_start = base_offset

    for sentence in sentences:
        if not current:
            current = sentence
        elif len(current) + len(sentence) + 1 <= chunk_size:
            current = current + " " + sentence
        else:
            chunks.append({
                "text": current,
                "chunk_index": 0,
                "start_offset": current_start,
                "end_offset": current_start + len(current),
            })
            current_start = current_start + len(current) + 1
            current = sentence

    if current.strip():
        chunks.append({
            "text": current,
            "chunk_index": 0,
            "start_offset": current_start,
            "end_offset": current_start + len(current),
        })

    return chunks
